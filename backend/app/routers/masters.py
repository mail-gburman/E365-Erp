
from datetime import datetime, date
from io import BytesIO
import threading, uuid, re
from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from openpyxl import load_workbook
from ..database import get_db, SessionLocal
from ..auth import get_current_user
from ..permissions import require_permission
from .. import models, schemas
from ..codegen import (
    next_vendor_code, next_client_code, next_product_code, next_asset_code,
    next_employee_code, next_po_code, next_po_number, next_equipment_master_code
)
from ..audit import audit

router = APIRouter(tags=["Masters"])

IMPORT_JOBS = {}

def _job_snapshot(job_id: str):
    job = IMPORT_JOBS.get(job_id)
    if not job:
        return None
    total = max(int(job.get("total_items", 0) or 0), 0)
    processed = max(int(job.get("processed_items", 0) or 0), 0)
    current_started = processed
    if total and job.get("current_item") and not job.get("done") and processed < total:
        current_started = min(processed + 1, total)
    pct = int((current_started / total) * 100) if total else 0
    if current_started > 0 and pct == 0:
        pct = 1
    return {
        **job,
        "current_index": current_started,
        "progress_pct": pct,
    }

def _set_job(job_id: str, **kwargs):
    if job_id not in IMPORT_JOBS:
        IMPORT_JOBS[job_id] = {}
    IMPORT_JOBS[job_id].update(kwargs)

def _iter_detected_inventory_rows(wb):
    for ws in wb.worksheets:
        header_row = _find_header_row(ws)
        norm = _build_norm_map(ws, header_row)
        if not norm:
            continue
        for r in range(header_row + 1, ws.max_row + 1):
            make = _gv(ws, norm, r, "make")
            model = _gv(ws, norm, r, "model")
            unique_no = _gv(ws, norm, r, "unique number")
            serial = _gv(ws, norm, r, "sl no")
            remarks = _gv(ws, norm, r, "remarks")
            if any([make, model, unique_no, serial, remarks]):
                yield ws, norm, r, make, model, unique_no, serial, remarks

def _iter_detected_generic_rows(wb, entity_type):
    ws = wb[wb.sheetnames[0]]
    header_row = _find_header_row(ws)
    norm = _build_norm_map(ws, header_row)
    for r in range(header_row + 1, ws.max_row + 1):
        if entity_type == "clients":
            name = _gv(ws, norm, r, "name")
        elif entity_type == "crew":
            name = _gv(ws, norm, r, "full name", "name")
        elif entity_type == "vendors":
            name = _gv(ws, norm, r, "name")
        else:
            name = _gv(ws, norm, r, "name")
        if name:
            yield ws, norm, r, name


def _clean(v):
    return str(v).strip() if v is not None else ""

def _parse_date(v):
    if v in (None, "", "None"): return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None

def _normalize_header(s):
    s = _clean(s).lower()
    s = s.replace(".", "").replace("_", " ").replace("-", " ").replace("/", " ")
    s = " ".join(s.split())
    return s

HEADER_ALIASES = {
    "unique number": ["unique no", "unique id", "asset code", "asset id", "code", "unique"],
    "make": ["brand", "manufacturer"],
    "model": ["model no", "model number"],
    "sl no": ["serial", "serial no", "slno", "s l no", "sr no", "serial number"],
    "qty": ["quantity", "qty nos"],
    "purchase date": ["purchasedate", "date of purchase", "purchase dt"],
    "remarks": ["remark", "comments", "comment", "note", "notes"],
    "name": ["client name", "vendor name", "warehouse name"],
    "full name": ["name", "employee name", "crew name", "staff name"],
    "role": ["designation"],
    "client code": ["code"],
    "vendor code": ["code"],
    "employee code": ["code", "staff code"],
    "vendor type": ["type"],
    "industry type": ["industry"],
    "billing address": ["address"],
    "gst number": ["gst", "gstin"],
    "contact person": ["contact name"],
    "manager name": ["manager"],
    "contact no": ["phone", "mobile"],
}

def _build_norm_map(ws, row_no):
    headers = [_normalize_header(ws.cell(row_no, c).value) for c in range(1, ws.max_column + 1)]
    return {h: i + 1 for i, h in enumerate(headers) if h}

def _find_header_row(ws):
    best_row, best_score = 1, -1
    expected = ["unique number", "make", "model", "sl no", "qty", "purchase date", "remarks", "name", "full name", "role"]
    for r in range(1, min(ws.max_row, 15) + 1):
        row = [_normalize_header(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        score = sum(1 for token in expected if token in row)
        if score > best_score:
            best_row, best_score = r, score
    return best_row

def _col_exists(norm, key):
    if key in norm:
        return True
    for alias in HEADER_ALIASES.get(key, []):
        if _normalize_header(alias) in norm:
            return True
    return False

def _gv(ws, norm, row_no, *keys):
    for key in keys:
        nk = _normalize_header(key)
        if nk in norm:
            return _clean(ws.cell(row_no, norm[nk]).value)
        for alias in HEADER_ALIASES.get(nk, []):
            alias_n = _normalize_header(alias)
            if alias_n in norm:
                return _clean(ws.cell(row_no, norm[alias_n]).value)
    return ""


def _max_prefixed_numeric(db: Session, model, field_name: str, prefix: str) -> int:
    col = getattr(model, field_name)
    vals = db.query(col).all()
    max_n = 0
    pat = re.compile(r"^" + re.escape(prefix) + r"(\d+)$")
    for (val,) in vals:
        if not val:
            continue
        m = pat.match(str(val))
        if m:
            max_n = max(max_n, int(m.group(1)))
    return max_n

def _next_prefixed_code_from_counter(prefix: str, width: int, current_value: int) -> tuple[str, int]:
    current_value += 1
    return f"{prefix}{current_value:0{width}d}", current_value

def _safe_product_code(db: Session):
    # Flush-based monotonic generation to avoid duplicate product_code within one import transaction
    while True:
        code = next_product_code(db)
        exists = db.query(models.InventoryItem).filter(models.InventoryItem.product_code == code).first()
        if not exists:
            return code
        db.flush()

def _infer_item_type(sheet_name: str):
    s = sheet_name.strip().lower()
    if s in ["lenses", "lens", "headset", "beltpack", "communication", "battery", "batteries", "charger", "audio", "audio accessories"]:
        return "accessory"
    return "device"

@router.get("/warehouses", response_model=list[schemas.WarehouseRead], dependencies=[Depends(require_permission("masters","view"))])
def list_warehouses(db: Session = Depends(get_db)):
    return db.query(models.Warehouse).order_by(models.Warehouse.name.asc()).all()

@router.post("/warehouses", response_model=schemas.WarehouseRead, dependencies=[Depends(require_permission("masters","add"))])
def create_warehouse(payload: schemas.WarehouseCreate, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    item = models.Warehouse(**payload.model_dump())
    db.add(item)
    audit(db, current_user.username, "create", "warehouse", details=payload.model_dump())
    db.commit(); db.refresh(item)
    return item

@router.put("/warehouses/{warehouse_id}", response_model=schemas.WarehouseRead, dependencies=[Depends(require_permission("masters", "edit"))])
def update_warehouse(warehouse_id: int, payload: schemas.WarehouseUpdate, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    item = db.query(models.Warehouse).filter(models.Warehouse.id == warehouse_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Warehouse not found.")
    data = payload.model_dump(exclude_unset=True)
    for k, v in data.items():
        setattr(item, k, v)
    audit(db, current_user.username, "update", "warehouse", entity_id=warehouse_id, details=data)
    db.commit(); db.refresh(item)
    return item

@router.get("/vendors", response_model=list[schemas.VendorRead], dependencies=[Depends(require_permission("masters","view"))])
def list_vendors(db: Session = Depends(get_db)):
    return db.query(models.Vendor).order_by(models.Vendor.name.asc()).all()

@router.post("/vendors", response_model=schemas.VendorRead, dependencies=[Depends(require_permission("masters","add"))])
def create_vendor(payload: schemas.VendorCreate, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    data = payload.model_dump()
    if not data.get("vendor_code"):
        data["vendor_code"] = next_vendor_code(db)
    item = models.Vendor(**data)
    db.add(item)
    audit(db, current_user.username, "create", "vendor", details=data)
    db.commit(); db.refresh(item)
    return item

@router.put("/vendors/{vendor_id}", response_model=schemas.VendorRead, dependencies=[Depends(require_permission("masters", "edit"))])
def update_vendor(vendor_id: int, payload: schemas.VendorUpdate, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    item = db.query(models.Vendor).filter(models.Vendor.id == vendor_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Vendor not found.")
    data = payload.model_dump(exclude_unset=True)
    for k, v in data.items():
        setattr(item, k, v)
    audit(db, current_user.username, "update", "vendor", entity_id=vendor_id, details=data)
    db.commit(); db.refresh(item)
    return item

@router.get("/clients", response_model=list[schemas.ClientRead], dependencies=[Depends(require_permission("masters","view"))])
def list_clients(db: Session = Depends(get_db)):
    return db.query(models.Client).order_by(models.Client.name.asc()).all()

@router.post("/clients", response_model=schemas.ClientRead, dependencies=[Depends(require_permission("masters","add"))])
def create_client(payload: schemas.ClientCreate, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    data = payload.model_dump()
    contacts = data.pop("contacts", [])
    primary_count = sum(1 for c in contacts if c.get("is_primary"))
    if contacts and primary_count != 1:
        raise HTTPException(status_code=400, detail="Exactly one primary contact is required.")
    item = models.Client(client_code=next_client_code(db), **data)
    db.add(item); db.commit(); db.refresh(item)
    for c in contacts:
        db.add(models.ClientContact(client_id=item.id, **c))
    audit(db, current_user.username, "create", "client", entity_id=item.id, details={"name": item.name, "contacts": contacts})
    db.commit(); db.refresh(item)
    return item

@router.put("/clients/{client_id}", response_model=schemas.ClientRead, dependencies=[Depends(require_permission("masters", "edit"))])
def update_client(client_id: int, payload: schemas.ClientUpdate, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    item = db.query(models.Client).filter(models.Client.id == client_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Client not found.")
    data = payload.model_dump(exclude_unset=True)
    for k, v in data.items():
        setattr(item, k, v)
    audit(db, current_user.username, "update", "client", entity_id=client_id, details=data)
    db.commit(); db.refresh(item)
    return item

@router.get("/equipment-master", response_model=list[schemas.EquipmentMasterRead], dependencies=[Depends(require_permission("masters","view"))])
def list_equipment_master(db: Session = Depends(get_db)):
    return db.query(models.EquipmentMaster).order_by(models.EquipmentMaster.name.asc()).all()

@router.post("/equipment-master", response_model=schemas.EquipmentMasterRead, dependencies=[Depends(require_permission("masters","add"))])
def create_equipment_master(payload: schemas.EquipmentMasterCreate, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    data = payload.model_dump()
    if not data.get("equipment_code"):
        data["equipment_code"] = next_equipment_master_code(db)
    item = models.EquipmentMaster(**data)
    db.add(item)
    audit(db, current_user.username, "create", "equipment_master", details=data)
    db.commit(); db.refresh(item)
    return item

@router.put("/equipment-master/{em_id}", response_model=schemas.EquipmentMasterRead, dependencies=[Depends(require_permission("masters", "edit"))])
def update_equipment_master(em_id: int, payload: schemas.EquipmentMasterUpdate, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    item = db.query(models.EquipmentMaster).filter(models.EquipmentMaster.id == em_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Equipment master not found.")
    data = payload.model_dump(exclude_unset=True)
    for k, v in data.items():
        setattr(item, k, v)
    audit(db, current_user.username, "update", "equipment_master", entity_id=em_id, details=data)
    db.commit(); db.refresh(item)
    return item

@router.get("/inventory", response_model=list[schemas.InventoryRead], dependencies=[Depends(require_permission("masters","view"))])
def list_inventory(db: Session = Depends(get_db)):
    return db.query(models.InventoryItem).order_by(models.InventoryItem.id.desc()).all()

@router.post("/inventory", response_model=schemas.InventoryRead, dependencies=[Depends(require_permission("masters","add"))])
def create_inventory(payload: schemas.InventoryCreate, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    data = payload.model_dump()
    if not data.get("asset_code"):
        data["asset_code"] = next_asset_code(db)
    item = models.InventoryItem(product_code=_safe_product_code(db), **data)
    db.add(item)
    db.flush()
    audit(db, current_user.username, "create", "inventory", details=data)
    db.commit(); db.refresh(item)
    return item

@router.put("/inventory/{item_id}", response_model=schemas.InventoryRead, dependencies=[Depends(require_permission("masters", "edit"))])
def update_inventory(item_id: int, payload: schemas.InventoryUpdate, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    item = db.query(models.InventoryItem).filter(models.InventoryItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Inventory item not found.")
    data = payload.model_dump(exclude_unset=True)
    for k, v in data.items():
        setattr(item, k, v)
    audit(db, current_user.username, "update", "inventory", entity_id=item_id, details=data)
    db.commit(); db.refresh(item)
    return item

@router.get("/crew", response_model=list[schemas.CrewRead], dependencies=[Depends(require_permission("masters","view"))])
def list_crew(db: Session = Depends(get_db)):
    return db.query(models.CrewMember).order_by(models.CrewMember.id.desc()).all()

@router.post("/crew", response_model=schemas.CrewRead, dependencies=[Depends(require_permission("masters","add"))])
def create_crew(payload: schemas.CrewCreate, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    import json as _json
    data = payload.model_dump()
    if not data.get("employee_code"):
        data["employee_code"] = next_employee_code(db)
    id_proofs = data.pop("id_proofs", None)
    if id_proofs:
        data["id_proofs_json"] = _json.dumps(id_proofs, default=str)
    item = models.CrewMember(**data)
    db.add(item)
    audit(db, current_user.username, "create", "crew", details=data)
    db.commit(); db.refresh(item)
    return item

@router.put("/crew/{crew_id}", response_model=schemas.CrewRead, dependencies=[Depends(require_permission("masters", "edit"))])
def update_crew(crew_id: int, payload: schemas.CrewUpdate, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    item = db.query(models.CrewMember).filter(models.CrewMember.id == crew_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Crew member not found.")
    data = payload.model_dump(exclude_unset=True)
    # Validate: external/contractual crew should have vendor_id
    manpower_type = data.get("manpower_type", item.manpower_type)
    vendor_id = data.get("vendor_id", item.vendor_id)
    if manpower_type in ("external", "contractual") and not vendor_id:
        raise HTTPException(status_code=400, detail="External/contractual crew must have a vendor (third-party supplier) assigned.")
    import json as _json
    if "id_proofs" in data:
        proofs = data.pop("id_proofs")
        item.id_proofs_json = _json.dumps(proofs, default=str) if proofs else None
    for k, v in data.items():
        setattr(item, k, v)
    audit(db, current_user.username, "update", "crew", entity_id=crew_id, details=data)
    db.commit(); db.refresh(item)
    return item

@router.get("/procurement", response_model=list[schemas.ProcurementRead], dependencies=[Depends(require_permission("masters","view"))])
def list_procurement(db: Session = Depends(get_db)):
    return db.query(models.ProcurementOrder).order_by(models.ProcurementOrder.id.desc()).all()

@router.post("/procurement", response_model=schemas.ProcurementRead, dependencies=[Depends(require_permission("masters","add"))])
def create_procurement(payload: schemas.ProcurementCreate, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    data = payload.model_dump()
    if not data.get("po_number"):
        data["po_number"] = next_po_number(db)
    item = models.ProcurementOrder(procurement_code=next_po_code(db), **data)
    db.add(item)
    audit(db, current_user.username, "create", "procurement", details=data)
    db.commit(); db.refresh(item)
    return item



@router.post("/bulk-upload-start", dependencies=[Depends(require_permission("masters","add"))])
async def bulk_upload_start(
    entity_type: str = Form(...),
    file: UploadFile = File(...),
    current_user = Depends(get_current_user),
):
    content = await file.read()
    job_id = str(uuid.uuid4())
    IMPORT_JOBS[job_id] = {
        "job_id": job_id,
        "state": "queued",
        "entity_type": entity_type,
        "total_items": 0,
        "processed_items": 0,
        "accepted_count": 0,
        "rejected_count": 0,
        "current_item": "",
        "current_sheet": "",
        "current_row": None,
        "accepted": [],
        "rejected": [],
        "message": "Queued for import",
        "done": False,
        "ok": None,
    }

    def worker():
        db = SessionLocal()
        try:
            try:
                wb = load_workbook(filename=BytesIO(content), data_only=True)
            except Exception as e:
                _set_job(job_id, state="failed", done=True, ok=False, message=f"Could not read workbook: {str(e)}")
                return

            if entity_type == "inventory":
                detected = list(_iter_detected_inventory_rows(wb))
                _set_job(job_id, state="running", total_items=len(detected), message=f"Detected {len(detected)} rows for import")
                product_seq = _max_prefixed_numeric(db, models.InventoryItem, "product_code", "PRD-")
                eqm_seq = _max_prefixed_numeric(db, models.EquipmentMaster, "equipment_code", "EQM-")
                for idx, row in enumerate(detected, start=1):
                    ws, norm, r, make, model, unique_no, serial, remarks = row
                    display_name = f"{make} {model}".strip() or unique_no or serial or f"Row {r}"
                    _set_job(job_id, processed_items=idx-1, current_item=display_name, current_sheet=ws.title, current_row=r, message=f"Adding {display_name} ({idx} of {len(detected)})")
                    try:
                        if not unique_no:
                            IMPORT_JOBS[job_id]["rejected"].append({"sheet": ws.title, "row": r, "status": "rejected", "reason": "Missing mandatory column value: UNIQUE NUMBER. Please provide client code format / unique number for this row."})
                            IMPORT_JOBS[job_id]["rejected_count"] += 1
                            continue

                        asset_code = unique_no
                        if db.query(models.InventoryItem).filter(models.InventoryItem.asset_code == asset_code).first():
                            IMPORT_JOBS[job_id]["rejected"].append({"sheet": ws.title, "row": r, "status": "rejected", "reason": f"Duplicate code warning: {asset_code} already exists"})
                            IMPORT_JOBS[job_id]["rejected_count"] += 1
                            continue

                        sheet_category = ws.title.strip().upper()
                        default_item_type = _infer_item_type(ws.title)

                        with db.begin_nested():
                            em = db.query(models.EquipmentMaster).filter(models.EquipmentMaster.name == display_name).first()
                            if not em:
                                eqm_code, eqm_seq = _next_prefixed_code_from_counter("EQM-", 5, eqm_seq)
                                em = models.EquipmentMaster(
                                    equipment_code=eqm_code,
                                    name=display_name,
                                    category=sheet_category,
                                    item_type=default_item_type,
                                    brand=make or None,
                                    model_no=model or None,
                                    notes=f"Imported from workbook sheet {ws.title}",
                                )
                                db.add(em)
                                db.flush()

                            product_code, product_seq = _next_prefixed_code_from_counter("PRD-", 6, product_seq)
                            item = models.InventoryItem(
                                asset_code=asset_code,
                                product_code=product_code,
                                name=display_name,
                                category=sheet_category,
                                item_type=default_item_type,
                                equipment_master_id=em.id if em else None,
                                status="available",
                                statutory_tag=serial or asset_code,
                                notes=remarks or f"Imported from {ws.title}",
                                warranty_expiry=None,
                                service_due=None,
                                service_status="ok",
                            )
                            db.add(item)
                            db.flush()

                        IMPORT_JOBS[job_id]["accepted"].append({"sheet": ws.title, "row": r, "status": "accepted", "code": asset_code, "name": display_name})
                        IMPORT_JOBS[job_id]["accepted_count"] += 1
                    except IntegrityError as e:
                        IMPORT_JOBS[job_id]["rejected"].append({"sheet": ws.title, "row": r, "status": "rejected", "reason": f"Database constraint warning: {str(e.orig)}"})
                        IMPORT_JOBS[job_id]["rejected_count"] += 1
                    except Exception as e:
                        IMPORT_JOBS[job_id]["rejected"].append({"sheet": ws.title, "row": r, "status": "rejected", "reason": f"Import parser error: {str(e)}"})
                        IMPORT_JOBS[job_id]["rejected_count"] += 1
                    finally:
                        _set_job(job_id, processed_items=idx)

            else:
                detected = list(_iter_detected_generic_rows(wb, entity_type))
                _set_job(job_id, state="running", total_items=len(detected), message=f"Detected {len(detected)} rows for import")
                for idx, row in enumerate(detected, start=1):
                    ws, norm, r, name = row
                    _set_job(job_id, processed_items=idx-1, current_item=name, current_sheet=ws.title, current_row=r, message=f"Adding {name} ({idx} of {len(detected)})")
                    try:
                        with db.begin_nested():
                            if entity_type == "clients":
                                code = _gv(ws, norm, r, "client code") or next_client_code(db)
                                if db.query(models.Client).filter(models.Client.client_code == code).first():
                                    raise ValueError(f"Duplicate code warning: {code} already exists")
                                db.add(models.Client(client_code=code, name=name, industry_type=_gv(ws, norm, r, "industry type") or None, billing_address=_gv(ws, norm, r, "billing address") or None, gst_number=_gv(ws, norm, r, "gst number") or None, notes=_gv(ws, norm, r, "notes") or None))
                            elif entity_type == "crew":
                                code = _gv(ws, norm, r, "employee code") or next_employee_code(db)
                                if db.query(models.CrewMember).filter(models.CrewMember.employee_code == code).first():
                                    raise ValueError(f"Duplicate code warning: {code} already exists")
                                phone_val = _gv(ws, norm, r, "phone", "contact number", "contact no", "mobile") or None
                                address_val = _gv(ws, norm, r, "address") or None
                                aadhar_val = _gv(ws, norm, r, "aadhar number", "aadhar", "aadhaar") or None
                                db.add(models.CrewMember(employee_code=code, full_name=name, role=_gv(ws, norm, r, "role") or "Crew", manpower_type=_gv(ws, norm, r, "manpower type") or "inhouse", home_station=_gv(ws, norm, r, "home station") or "Base", phone=phone_val, address=address_val, aadhar_number=aadhar_val, status=_gv(ws, norm, r, "status") or "available"))
                            elif entity_type == "vendors":
                                code = _gv(ws, norm, r, "vendor code") or next_vendor_code(db)
                                if db.query(models.Vendor).filter(models.Vendor.vendor_code == code).first():
                                    raise ValueError(f"Duplicate code warning: {code} already exists")
                                db.add(models.Vendor(vendor_code=code, name=name, vendor_type=_gv(ws, norm, r, "vendor type", "type") or "equipment", city=_gv(ws, norm, r, "city") or None, contact_person=_gv(ws, norm, r, "contact person") or None, phone=_gv(ws, norm, r, "phone") or None, email=_gv(ws, norm, r, "email") or None, gst_number=_gv(ws, norm, r, "gst number") or None, notes=_gv(ws, norm, r, "notes") or None))
                            elif entity_type == "warehouses":
                                code = _gv(ws, norm, r, "code") or f"WH-{r}"
                                if db.query(models.Warehouse).filter(models.Warehouse.code == code).first():
                                    raise ValueError(f"Duplicate code warning: {code} already exists")
                                db.add(models.Warehouse(code=code, name=name, city=_gv(ws, norm, r, "city") or "City", address=_gv(ws, norm, r, "address") or None, manager_name=_gv(ws, norm, r, "manager name") or None, contact_no=_gv(ws, norm, r, "contact no") or None))
                            else:
                                raise ValueError("Unsupported entity type for bulk import")
                            db.flush()
                        IMPORT_JOBS[job_id]["accepted"].append({"sheet": ws.title, "row": r, "status": "accepted", "code": code, "name": name})
                        IMPORT_JOBS[job_id]["accepted_count"] += 1
                    except Exception as e:
                        IMPORT_JOBS[job_id]["rejected"].append({"sheet": ws.title, "row": r, "status": "rejected", "reason": str(e)})
                        IMPORT_JOBS[job_id]["rejected_count"] += 1
                    finally:
                        _set_job(job_id, processed_items=idx)

            audit(db, current_user.username, "bulk_upload", entity_type, details={"job_id": job_id, "accepted": IMPORT_JOBS[job_id]["accepted_count"], "rejected": IMPORT_JOBS[job_id]["rejected_count"]})
            db.commit()
            _set_job(job_id, state="completed", done=True, ok=True, message="Import completed.")
        except Exception as e:
            db.rollback()
            _set_job(job_id, state="failed", done=True, ok=False, message=f"Import failed: {str(e)}")
        finally:
            db.close()

    threading.Thread(target=worker, daemon=True).start()
    return {"ok": True, "job_id": job_id}

@router.get("/bulk-upload-status/{job_id}", dependencies=[Depends(require_permission("masters","add"))])
def bulk_upload_status(job_id: str):
    snap = _job_snapshot(job_id)
    if not snap:
        raise HTTPException(status_code=404, detail="Import job not found.")
    return snap


@router.post("/bulk-upload-preview", dependencies=[Depends(require_permission("masters","add"))])
async def bulk_upload_preview(
    entity_type: str = Form(...),
    file: UploadFile = File(...),
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    content = await file.read()
    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as e:
        return {"ok": False, "detected_total": 0, "sheets": [], "sample_items": [], "reason": f"Could not read workbook: {str(e)}"}

    sheets = []
    sample_items = []
    detected_total = 0

    if entity_type == "inventory":
        for ws in wb.worksheets:
            header_row = _find_header_row(ws)
            norm = _build_norm_map(ws, header_row)
            detected = 0
            for r in range(header_row + 1, ws.max_row + 1):
                make = _gv(ws, norm, r, "make")
                model = _gv(ws, norm, r, "model")
                unique_no = _gv(ws, norm, r, "unique number")
                serial = _gv(ws, norm, r, "sl no")
                remarks = _gv(ws, norm, r, "remarks")
                if not any([make, model, unique_no, serial, remarks]):
                    continue
                detected += 1
                name = f"{make} {model}".strip() or unique_no or serial or f"Row {r}"
                if len(sample_items) < 20:
                    sample_items.append({"sheet": ws.title, "row": r, "name": name, "code": unique_no or "-"})
            sheets.append({"sheet": ws.title, "detected_rows": detected})
            detected_total += detected
        return {"ok": True, "detected_total": detected_total, "sheets": sheets, "sample_items": sample_items}

    ws = wb[wb.sheetnames[0]]
    header_row = _find_header_row(ws)
    norm = _build_norm_map(ws, header_row)
    detected = 0
    for r in range(header_row + 1, ws.max_row + 1):
        if entity_type == "clients":
            name = _gv(ws, norm, r, "name")
        elif entity_type == "crew":
            name = _gv(ws, norm, r, "full name", "name")
        elif entity_type == "vendors":
            name = _gv(ws, norm, r, "name")
        else:
            name = _gv(ws, norm, r, "name")
        if not name:
            continue
        detected += 1
        if len(sample_items) < 20:
            sample_items.append({"sheet": ws.title, "row": r, "name": name, "code": "-"})
    return {"ok": True, "detected_total": detected, "sheets": [{"sheet": ws.title, "detected_rows": detected}], "sample_items": sample_items}


@router.post("/bulk-upload", dependencies=[Depends(require_permission("masters","add"))])
async def bulk_upload(
    entity_type: str = Form(...),
    file: UploadFile = File(...),
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    content = await file.read()
    accepted = []
    rejected = []
    count = 0

    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as e:
        return {"ok": False, "count": 0, "accepted": [], "rejected": [{"sheet": "-", "row": "-", "status": "rejected", "reason": f"Could not read workbook: {str(e)}"}]}

    if entity_type == "inventory":
        for ws in wb.worksheets:
            header_row = _find_header_row(ws)
            norm = _build_norm_map(ws, header_row)
            if not norm:
                rejected.append({"sheet": ws.title, "row": header_row, "status": "rejected", "reason": "Missing mandatory column warning: could not detect header row"})
                continue

            for r in range(header_row + 1, ws.max_row + 1):
                make = _gv(ws, norm, r, "make")
                model = _gv(ws, norm, r, "model")
                unique_no = _gv(ws, norm, r, "unique number")
                serial = _gv(ws, norm, r, "sl no")
                remarks = _gv(ws, norm, r, "remarks")

                if not any([make, model, unique_no, serial, remarks]):
                    continue

                if not unique_no:
                    rejected.append({
                        "sheet": ws.title,
                        "row": r,
                        "status": "rejected",
                        "reason": "Missing mandatory column value: UNIQUE NUMBER. Please provide client code format / unique number for this row."
                    })
                    continue

                asset_code = unique_no
                if db.query(models.InventoryItem).filter(models.InventoryItem.asset_code == asset_code).first():
                    rejected.append({"sheet": ws.title, "row": r, "status": "rejected", "reason": f"Duplicate code warning: {asset_code} already exists"})
                    continue

                display_name = f"{make} {model}".strip() or asset_code
                sheet_category = ws.title.strip().upper()
                default_item_type = _infer_item_type(ws.title)

                try:
                    with db.begin_nested():
                        em = db.query(models.EquipmentMaster).filter(models.EquipmentMaster.name == display_name).first()
                        if not em:
                            em = models.EquipmentMaster(
                                equipment_code=next_equipment_master_code(db),
                                name=display_name,
                                category=sheet_category,
                                item_type=default_item_type,
                                brand=make or None,
                                model_no=model or None,
                                notes=f"Imported from workbook sheet {ws.title}",
                            )
                            db.add(em)
                            db.flush()

                        item = models.InventoryItem(
                            asset_code=asset_code,
                            product_code=_safe_product_code(db),
                            name=display_name,
                            category=sheet_category,
                            item_type=default_item_type,
                            equipment_master_id=em.id if em else None,
                            status="available",
                            statutory_tag=serial or asset_code,
                            notes=remarks or f"Imported from {ws.title}",
                            warranty_expiry=None,
                            service_due=None,
                            service_status="ok",
                        )
                        db.add(item)
                        db.flush()

                    count += 1
                    accepted.append({"sheet": ws.title, "row": r, "status": "accepted", "code": asset_code, "name": display_name})
                except IntegrityError as e:
                    rejected.append({"sheet": ws.title, "row": r, "status": "rejected", "reason": f"Database constraint warning: {str(e.orig)}"})
                except Exception as e:
                    rejected.append({"sheet": ws.title, "row": r, "status": "rejected", "reason": f"Import parser error: {str(e)}"})

        audit(db, current_user.username, "bulk_upload", "inventory", details={"count": count, "filename": file.filename, "accepted": len(accepted), "rejected": len(rejected)})
        db.commit()
        return {"ok": True, "count": count, "accepted": accepted, "rejected": rejected}

    ws = wb[wb.sheetnames[0]]
    header_row = _find_header_row(ws)
    norm = _build_norm_map(ws, header_row)

    required_cols = {
        "clients": ["name"],
        "crew": ["full name"],
        "vendors": ["name"],
        "warehouses": ["name"],
    }.get(entity_type, [])

    missing = [c for c in required_cols if not _col_exists(norm, c)]
    if missing:
        return {"ok": False, "count": 0, "accepted": [], "rejected": [{"sheet": ws.title, "row": header_row, "status": "rejected", "reason": f"Missing mandatory column warning: {', '.join(missing)}"}]}

    for r in range(header_row + 1, ws.max_row + 1):
        try:
            with db.begin_nested():
                if entity_type == "clients":
                    name = _gv(ws, norm, r, "name")
                    if not name:
                        rejected.append({"sheet": ws.title, "row": r, "status": "rejected", "reason": "Missing mandatory column value: name"})
                        continue
                    code = _gv(ws, norm, r, "client code") or next_client_code(db)
                    if db.query(models.Client).filter(models.Client.client_code == code).first():
                        rejected.append({"sheet": ws.title, "row": r, "status": "rejected", "reason": f"Duplicate code warning: {code} already exists"})
                        continue
                    db.add(models.Client(client_code=code, name=name, industry_type=_gv(ws, norm, r, "industry type") or None, billing_address=_gv(ws, norm, r, "billing address") or None, gst_number=_gv(ws, norm, r, "gst number") or None, notes=_gv(ws, norm, r, "notes") or None))
                    accepted.append({"sheet": ws.title, "row": r, "status": "accepted", "code": code, "name": name}); count += 1

                elif entity_type == "crew":
                    name = _gv(ws, norm, r, "full name", "name")
                    if not name:
                        rejected.append({"sheet": ws.title, "row": r, "status": "rejected", "reason": "Missing mandatory column value: full_name"})
                        continue
                    code = _gv(ws, norm, r, "employee code") or next_employee_code(db)
                    if db.query(models.CrewMember).filter(models.CrewMember.employee_code == code).first():
                        rejected.append({"sheet": ws.title, "row": r, "status": "rejected", "reason": f"Duplicate code warning: {code} already exists"})
                        continue
                    phone_val = _gv(ws, norm, r, "phone", "contact number", "contact no", "mobile") or None
                    address_val = _gv(ws, norm, r, "address") or None
                    aadhar_val = _gv(ws, norm, r, "aadhar number", "aadhar", "aadhaar") or None
                    db.add(models.CrewMember(employee_code=code, full_name=name, role=_gv(ws, norm, r, "role") or "Crew", manpower_type=_gv(ws, norm, r, "manpower type") or "inhouse", home_station=_gv(ws, norm, r, "home station") or "Base", phone=phone_val, address=address_val, aadhar_number=aadhar_val, status=_gv(ws, norm, r, "status") or "available"))
                    accepted.append({"sheet": ws.title, "row": r, "status": "accepted", "code": code, "name": name}); count += 1

                elif entity_type == "vendors":
                    name = _gv(ws, norm, r, "name")
                    if not name:
                        rejected.append({"sheet": ws.title, "row": r, "status": "rejected", "reason": "Missing mandatory column value: name"})
                        continue
                    code = _gv(ws, norm, r, "vendor code") or next_vendor_code(db)
                    if db.query(models.Vendor).filter(models.Vendor.vendor_code == code).first():
                        rejected.append({"sheet": ws.title, "row": r, "status": "rejected", "reason": f"Duplicate code warning: {code} already exists"})
                        continue
                    db.add(models.Vendor(vendor_code=code, name=name, vendor_type=_gv(ws, norm, r, "vendor type", "type") or "equipment", city=_gv(ws, norm, r, "city") or None, contact_person=_gv(ws, norm, r, "contact person") or None, phone=_gv(ws, norm, r, "phone") or None, email=_gv(ws, norm, r, "email") or None, gst_number=_gv(ws, norm, r, "gst number") or None, notes=_gv(ws, norm, r, "notes") or None))
                    accepted.append({"sheet": ws.title, "row": r, "status": "accepted", "code": code, "name": name}); count += 1

                elif entity_type == "warehouses":
                    name = _gv(ws, norm, r, "name")
                    if not name:
                        rejected.append({"sheet": ws.title, "row": r, "status": "rejected", "reason": "Missing mandatory column value: name"})
                        continue
                    code = _gv(ws, norm, r, "code") or f"WH-{r}"
                    if db.query(models.Warehouse).filter(models.Warehouse.code == code).first():
                        rejected.append({"sheet": ws.title, "row": r, "status": "rejected", "reason": f"Duplicate code warning: {code} already exists"})
                        continue
                    db.add(models.Warehouse(code=code, name=name, city=_gv(ws, norm, r, "city") or "City", address=_gv(ws, norm, r, "address") or None, manager_name=_gv(ws, norm, r, "manager name") or None, contact_no=_gv(ws, norm, r, "contact no") or None))
                    accepted.append({"sheet": ws.title, "row": r, "status": "accepted", "code": code, "name": name}); count += 1

        except IntegrityError as e:
            rejected.append({"sheet": ws.title, "row": r, "status": "rejected", "reason": f"Database constraint warning: {str(e.orig)}"})
        except Exception as e:
            rejected.append({"sheet": ws.title, "row": r, "status": "rejected", "reason": f"Import parser error: {str(e)}"})

    audit(db, current_user.username, "bulk_upload", entity_type, details={"count": count, "filename": file.filename, "accepted": len(accepted), "rejected": len(rejected)})
    db.commit()
    return {"ok": True, "count": count, "accepted": accepted, "rejected": rejected}


# ── Smart Upload (Universal Parser) ──

from ..universal_parser import smart_preview, apply_mapping_and_import, ENTITY_SCHEMAS
import json as _json

@router.post("/smart-upload/preview", dependencies=[Depends(require_permission("masters","add"))])
async def smart_upload_preview(
    file: UploadFile = File(...),
    entity_type: str = Form(None),
):
    content = await file.read()
    try:
        result = smart_preview(file.filename, content, entity_type or None)
        return {"ok": True, **result}
    except ValueError as e:
        return {"ok": False, "error": str(e)}
    except Exception as e:
        return {"ok": False, "error": f"Could not parse file: {str(e)}"}


@router.post("/smart-upload/import", dependencies=[Depends(require_permission("masters","add"))])
async def smart_upload_import(
    file: UploadFile = File(...),
    entity_type: str = Form(...),
    mapping: str = Form(...),
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    content = await file.read()
    try:
        user_mapping = _json.loads(mapping)  # { field_name: col_index }
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid mapping JSON.")
    try:
        result = apply_mapping_and_import(file.filename, content, entity_type, user_mapping, db)
        audit(db, current_user.username, "smart_upload", entity_type, details={"filename": file.filename, "count": result["count"], "accepted": len(result["accepted"]), "rejected": len(result["rejected"])})
        db.commit()
        return {"ok": True, **result}
    except ValueError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")


# ─── STATUTORY ID VERIFICATION ───────────────────────────────────────────────

_INDIAN_STATES = {
    "01":"Jammu & Kashmir","02":"Himachal Pradesh","03":"Punjab","04":"Chandigarh",
    "05":"Uttarakhand","06":"Haryana","07":"Delhi","08":"Rajasthan","09":"Uttar Pradesh",
    "10":"Bihar","11":"Sikkim","12":"Arunachal Pradesh","13":"Nagaland","14":"Manipur",
    "15":"Mizoram","16":"Tripura","17":"Meghalaya","18":"Assam","19":"West Bengal",
    "20":"Jharkhand","21":"Odisha","22":"Chhattisgarh","23":"Madhya Pradesh",
    "24":"Gujarat","26":"Dadra & NH and D&D","27":"Maharashtra","28":"Andhra Pradesh",
    "29":"Karnataka","30":"Goa","31":"Lakshadweep","32":"Kerala","33":"Tamil Nadu",
    "34":"Puducherry","35":"A&N Islands","36":"Telangana","37":"Andhra Pradesh (New)",
    "38":"Ladakh","97":"Other Territory","99":"Centre Jurisdiction",
}

_GSTIN_CHARS = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"

def _gstin_checksum(gstin14: str) -> str:
    total = 0
    for i, ch in enumerate(gstin14.upper()):
        code = _GSTIN_CHARS.index(ch)
        mult = 1 if i % 2 == 0 else 2
        prod = code * mult
        total += prod // 36 + prod % 36
    return _GSTIN_CHARS[(36 - total % 36) % 36]

# Verhoeff tables
_V_D = [[0,1,2,3,4,5,6,7,8,9],[1,2,3,4,0,6,7,8,9,5],[2,3,4,0,1,7,8,9,5,6],
        [3,4,0,1,2,8,9,5,6,7],[4,0,1,2,3,9,5,6,7,8],[5,9,8,7,6,0,4,3,2,1],
        [6,5,9,8,7,1,0,4,3,2],[7,6,5,9,8,2,1,0,4,3],[8,7,6,5,9,3,2,1,0,4],
        [9,8,7,6,5,4,3,2,1,0]]
_V_P = [[0,1,2,3,4,5,6,7,8,9],[1,5,7,6,2,8,3,0,9,4],[5,8,0,3,7,9,6,1,4,2],
        [8,9,1,6,0,4,3,5,2,7],[9,4,5,3,1,2,6,8,7,0],[4,2,8,6,5,7,3,9,0,1],
        [2,7,9,3,8,0,6,4,1,5],[7,0,4,6,9,1,3,2,5,8]]

def _aadhaar_verhoeff(num: str) -> bool:
    digits = [int(d) for d in reversed(num)]
    c = 0
    for i, d in enumerate(digits):
        c = _V_D[c][_V_P[i % 8][d]]
    return c == 0


@router.get("/verify-id/{id_type}", dependencies=[Depends(get_current_user)])
def verify_statutory_id(id_type: str, value: str):
    """
    Verify format + checksum of Indian statutory IDs.
    Returns { valid, checksum_valid, message, info }.
    """
    v = value.strip().replace(" ", "").upper()
    t = id_type.strip()

    PATTERNS = {
        "Aadhaar":          r"^[2-9]\d{11}$",
        "PAN":              r"^[A-Z]{5}[0-9]{4}[A-Z]$",
        "Passport":         r"^[A-Z][1-9]\d{6}$",
        "Driving License":  r"^[A-Z]{2}\d{2}[\s-]?\d{4}\d{7}$",
        "Voter ID":         r"^[A-Z]{3}\d{7}$",
        "GST":              r"^\d{2}[A-Z]{5}\d{4}[A-Z][1-9A-Z]Z[0-9A-Z]$",
    }

    if t == "Others":
        return {"valid": True, "message": "No validation for Others type"}

    pat = PATTERNS.get(t)
    if not pat:
        raise HTTPException(status_code=400, detail=f"Unknown ID type: {t}")

    if not re.match(pat, v):
        hints = {
            "Aadhaar": "12 digits starting with 2–9",
            "PAN": "AAAAA9999A (5 letters, 4 digits, 1 letter)",
            "Passport": "A1234567 (letter + 7 digits)",
            "Driving License": "MH01 2018 1234567",
            "Voter ID": "ABC1234567 (3 letters + 7 digits)",
            "GST": "22AAAAA0000A1Z5 (15 characters)",
        }
        return {"valid": False, "checksum_valid": None,
                "message": f"Invalid format — {hints.get(t, 'check format')}"}

    # Checksum
    if t == "GST":
        expected = _gstin_checksum(v[:14])
        ok = expected == v[14]
        state_code = v[:2]
        embedded_pan = v[2:12]
        state = _INDIAN_STATES.get(state_code, f"State {state_code}")
        return {
            "valid": ok,
            "checksum_valid": ok,
            "message": f"GSTIN checksum {'valid' if ok else 'INVALID (expected ' + expected + ')'}",
            "info": {"state": state, "state_code": state_code, "embedded_pan": embedded_pan} if ok else None,
        }

    if t == "Aadhaar":
        ok = _aadhaar_verhoeff(v)
        return {
            "valid": ok,
            "checksum_valid": ok,
            "message": "Aadhaar checksum valid" if ok else "Aadhaar checksum invalid — number may be incorrect",
        }

    return {"valid": True, "checksum_valid": None, "message": f"{t} format valid"}
